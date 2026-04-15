import pymysql
import pandas as pd
import smtplib
import json
import io
import re
import unicodedata
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from datetime import date, timedelta
import logging

from openpyxl.chart import BarChart, Reference


logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(message)s")
logger = logging.getLogger(__name__)


# --- CARREGAR CONFIGURAÇÃO (usa o MESMO config.json atual, só para MySQL) ---
def load_config(path: str = "config.json") -> dict:
    try:
        with open(path, "r") as f:
            return json.load(f)
    except Exception as e:
        print(f"Erro ao ler {path}: {e}")
        raise e


config = load_config()


# CONFIGURAÇÕES DE EMAIL (SES)
SMTP_SERVER = "email-smtp.us-east-1.amazonaws.com"
SMTP_PORT = 587
SMTP_USER = "AKIAXWX65XLPPLLZNEYV"
SMTP_PASS = "BLlfuFIkyN+aONFwCI6axjfLJYvVgFd0iRIk9n3LlElH"
EMAIL_REMETENTE = "nao-responda@inhousemarket.com.br"

# E-mails padrão de cópia (CC) e cópia oculta (CCO)
EMAILS_COPIA_PADRAO = []
EMAILS_CCO_PADRAO = [
    "caiquec93@gmail.com",
    "vitor.machado@inhousemarket.com.br",
    "cassio@inhousemarket.com.br",
]


def clean_filename(text: str) -> str:
    nfkd = unicodedata.normalize("NFKD", text)
    ascii_text = nfkd.encode("ASCII", "ignore").decode("utf-8")
    clean = re.sub(r"[^a-zA-Z0-9]", "_", ascii_text)
    return re.sub(r"_+", "_", clean).strip("_")


def get_mysql_conn():
    try:
        return pymysql.connect(
            host=config["mysql"]["host"],
            user=config["mysql"]["user"],
            password=config["mysql"]["password"],
            database=config["mysql"]["database"],
            connect_timeout=30,
        )
    except Exception as e:
        logger.error(f"Erro MySQL: {e}")
        raise e


def _parse_emails(valor):
    if not valor:
        return []
    if isinstance(valor, (list, tuple, set)):
        base = []
        for v in valor:
            if not v:
                continue
            base.extend(re.split(r"[;,]", str(v)))
    else:
        base = re.split(r"[;,]", str(valor))
    return [e.strip() for e in base if e and e.strip()]


def enviar_email(destinatario, assunto, html, anexo, emails_copia=None, emails_cco=None):
    """Envia e-mail com opção de Cópia (CC) e Cópia Oculta (CCO)."""

    # Trata destinatário principal (pode ser string ou lista)
    if isinstance(destinatario, (list, tuple, set)):
        to_list = _parse_emails(destinatario)
    else:
        to_list = _parse_emails(destinatario)

    cc_list = _parse_emails(emails_copia)
    bcc_list = _parse_emails(emails_cco)

    msg = MIMEMultipart()
    msg["From"] = EMAIL_REMETENTE
    if to_list:
        msg["To"] = ", ".join(to_list)
    if cc_list:
        msg["Cc"] = ", ".join(cc_list)
    msg["Subject"] = assunto
    msg.attach(MIMEText(html, "html"))

    part = MIMEBase("application", "octet-stream")
    part.set_payload(anexo["dados"].getvalue())
    encoders.encode_base64(part)
    part.add_header("Content-Disposition", "attachment", filename=anexo["nome"])
    msg.attach(part)

    # Lista final de destinatários SMTP (To + Cc + Bcc)
    destinatarios_smtp = []
    destinatarios_smtp.extend(to_list)
    destinatarios_smtp.extend(cc_list)
    destinatarios_smtp.extend(bcc_list)

    try:
        server = smtplib.SMTP(SMTP_SERVER, SMTP_PORT)
        server.starttls()
        server.login(SMTP_USER, SMTP_PASS)
        server.sendmail(EMAIL_REMETENTE, destinatarios_smtp, msg.as_string())
        server.quit()
        return True, "Enviado com sucesso!"
    except Exception as e:
        return False, str(e)


def gerar_corpo_html(parceiro, periodo, data_ini, data_fim):
    return f"""
    <html>
      <body style="font-family: Arial, sans-serif;">
        <h2 style="color: #0047AB;">Relatório Consolidado - {parceiro}</h2>
        <p>Olá, segue em anexo o arquivo contendo os dados de <b>todas as lojas</b>.</p>
        <p><b>Referência:</b> {periodo}</p>
        <p><b>Período:</b> {data_ini.strftime('%d/%m/%Y')} a {data_fim.strftime('%d/%m/%Y')}</p>
        <hr><p><small>Sistema InHouse</small></p>
      </body>
    </html>
    """


def gerar_arquivo_excel_consolidado(df_consolidado, parceiro_nome, periodo_atual, dt_ini_display, dt_fim_display):
    """Gera Excel com detalhe, resumo por loja, gráficos e formato BR."""
    buffer = io.BytesIO()

    if "Quantidade" in df_consolidado.columns:
        df_consolidado["Quantidade"] = pd.to_numeric(df_consolidado["Quantidade"], errors="coerce")
    if "Valor" in df_consolidado.columns:
        df_consolidado["Valor"] = pd.to_numeric(df_consolidado["Valor"], errors="coerce")

    pivot_loja = (
        df_consolidado
        .pivot_table(index="Loja", values=["Quantidade", "Valor"], aggfunc="sum")
        .reset_index()
    )

    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df_consolidado.to_excel(writer, sheet_name="Detalhe_Vendas", index=False)
        pivot_loja.to_excel(writer, sheet_name="Resumo_Por_Loja", index=False)

        wb = writer.book
        ws_detalhe = writer.sheets["Detalhe_Vendas"]
        ws_resumo = writer.sheets["Resumo_Por_Loja"]

        header_row = 1
        cols = {cell.value: cell.column for cell in ws_detalhe[header_row]}
        col_qtd = cols.get("Quantidade")
        col_val = cols.get("Valor")

        fmt_valor = "#,##0.00"
        fmt_qtd = "#,##0"

        for row in range(header_row + 1, ws_detalhe.max_row + 1):
            if col_qtd:
                c = ws_detalhe.cell(row=row, column=col_qtd)
                c.number_format = fmt_qtd
            if col_val:
                c = ws_detalhe.cell(row=row, column=col_val)
                c.number_format = fmt_valor

        header_row_resumo = 1
        cols_resumo = {cell.value: cell.column for cell in ws_resumo[header_row_resumo]}
        col_qtd_r = cols_resumo.get("Quantidade")
        col_val_r = cols_resumo.get("Valor")

        for row in range(header_row_resumo + 1, ws_resumo.max_row + 1):
            if col_qtd_r:
                c = ws_resumo.cell(row=row, column=col_qtd_r)
                c.number_format = fmt_qtd
            if col_val_r:
                c = ws_resumo.cell(row=row, column=col_val_r)
                c.number_format = fmt_valor

        max_row = ws_resumo.max_row

        # Gráfico 1: Valor por loja
        chart_valor = BarChart()
        chart_valor.title = "Faturamento por Loja"
        chart_valor.y_axis.title = "Valor"
        chart_valor.x_axis.title = "Loja"

        data_val = Reference(ws_resumo, min_col=col_val_r, max_col=col_val_r, min_row=1, max_row=max_row)
        cats_val = Reference(ws_resumo, min_col=cols_resumo["Loja"], max_col=cols_resumo["Loja"], min_row=2, max_row=max_row)

        chart_valor.add_data(data_val, titles_from_data=True)
        chart_valor.set_categories(cats_val)
        chart_valor.height = 10
        chart_valor.width = 24

        ws_resumo.add_chart(chart_valor, "F5")

        # Gráfico 2: Quantidade por loja
        chart_qtd = BarChart()
        chart_qtd.title = "Quantidade por Loja"
        chart_qtd.y_axis.title = "Quantidade"
        chart_qtd.x_axis.title = "Loja"

        data_qtd = Reference(ws_resumo, min_col=col_qtd_r, max_col=col_qtd_r, min_row=1, max_row=max_row)
        cats_qtd = Reference(ws_resumo, min_col=cols_resumo["Loja"], max_col=cols_resumo["Loja"], min_row=2, max_row=max_row)

        chart_qtd.add_data(data_qtd, titles_from_data=True)
        chart_qtd.set_categories(cats_qtd)
        chart_qtd.height = 10
        chart_qtd.width = 24

        ws_resumo.add_chart(chart_qtd, "F25")

    safe_parceiro = clean_filename(parceiro_nome)
    filename = f"Relatorio_Consolidado_{periodo_atual}_{safe_parceiro}.xlsx"

    anexo_unico = {"nome": filename, "dados": buffer}
    return anexo_unico


def obter_config_disparo(modos_ativos):
    """Retorna DataFrame com configuração de disparo filtrada por período.
    LIKE_PATTERN filtra lojas pelo nome no MySQL (st.name LIKE ...).
    Novas lojas do parceiro entram automaticamente, sem alterar o código.
    """
    dados_config = [
        # Luggo — todas as lojas cujo nome contém "luggo"
        {"PARCEIRO": "Luggo",  "LIKE_PATTERN": "%luggo%",  "EMAIL_DESTINO": "elizangela.zico@luggo.com.br",  "PERIODO": "MENSAL",  "ATIVO": True},

        # MRV — todas as lojas cujo nome contém "mrv"
        {"PARCEIRO": "MRV",    "LIKE_PATTERN": "%mrv%",    "EMAIL_DESTINO": "servicos@mrv.com.br",           "PERIODO": "MENSAL",  "ATIVO": True},

        # Selfit — todas as lojas cujo nome contém "selfit"
        {"PARCEIRO": "Selfit", "LIKE_PATTERN": "%selfit%", "EMAIL_DESTINO": "selfitparceria@inhouse.com.br", "PERIODO": "SEMANAL", "ATIVO": True},
        {"PARCEIRO": "Selfit", "LIKE_PATTERN": "%selfit%", "EMAIL_DESTINO": "selfitparceria@inhouse.com.br", "PERIODO": "MENSAL",  "ATIVO": True},
    ]

    df_config = pd.DataFrame(dados_config)
    df_config = df_config[(df_config["ATIVO"] == True) & (df_config["PERIODO"].isin(modos_ativos))]
    return df_config


def main():
    logger.info("--- INICIANDO ROBÔ CONSOLIDADO v2 (MYSQL) ---")
    hoje = date.today()

    # Calendário automático: segunda-feira dispara SEMANAL, dia 1 do mês dispara MENSAL.
    # Para forçar manualmente, mude para FORCAR_MODO = ["SEMANAL", "MENSAL"].
    FORCAR_MODO = []

    modos_ativos = list(FORCAR_MODO) if FORCAR_MODO else []
    if not FORCAR_MODO:
        if hoje.weekday() == 0:
            modos_ativos.append("SEMANAL")
        if hoje.day == 1:
            modos_ativos.append("MENSAL")

    if not modos_ativos:
        logger.info("Hoje não é dia de envio segundo as regras de calendário.")
        return

    logger.info(f"Modos ativos: {modos_ativos}")

    df_config = obter_config_disparo(modos_ativos)
    if df_config.empty:
        logger.info("Nenhuma configuração ativa para os modos atuais.")
        return

    logger.info(f"Conectando no MySQL em {config['mysql']['host']} ...")
    try:
        conn_mysql = get_mysql_conn()

        for _, row in df_config.iterrows():
            parceiro_nome = row["PARCEIRO"]
            periodo_atual = row["PERIODO"]
            like_pattern  = row["LIKE_PATTERN"]

            logger.info(f"> Processando {parceiro_nome} - {periodo_atual} (LIKE '{like_pattern}')...")

            if periodo_atual == "SEMANAL":
                dt_fim = hoje - timedelta(days=1)
                dt_ini = hoje - timedelta(days=7)
            else:  # MENSAL
                dt_fim = date(hoje.year, hoje.month, 1) - timedelta(days=1)
                dt_ini = date(dt_fim.year, dt_fim.month, 1)

            query_mysql = f"""
            SELECT
                DATE_FORMAT(si.created_at, '%d/%m/%Y %H:%i') AS `Data/Hora`,
                st.name                                       AS Loja,
                p.name                                        AS Produto,
                si.quantity                                   AS Quantidade,
                (si.price / 100.0)                           AS Valor
            FROM sales_items si
            INNER JOIN sales    s  ON si.sales_id   = s.id
            INNER JOIN stores   st ON s.store_id    = st.id
            INNER JOIN products p  ON si.product_id = p.id
            WHERE st.name LIKE '{like_pattern}'
              AND s.status = 'paid'
              AND si.created_at >= '{dt_ini} 00:00:00'
              AND si.created_at <= '{dt_fim} 23:59:59'
            ORDER BY st.name, si.created_at DESC
            """

            try:
                with conn_mysql.cursor() as cursor:
                    cursor.execute(query_mysql)
                    dados_brutos = cursor.fetchall()

                colunas = ["Data/Hora", "Loja", "Produto", "Quantidade", "Valor"]
                df_consolidado = pd.DataFrame(dados_brutos, columns=colunas)

                if df_consolidado.empty:
                    logger.info(f"  --> {parceiro_nome}: Sem vendas no período ({dt_ini} a {dt_fim}).")
                    continue

                lojas_encontradas = df_consolidado["Loja"].nunique()
                logger.info(f"  Lojas encontradas: {lojas_encontradas} | Linhas: {len(df_consolidado)}")

                anexo_unico = gerar_arquivo_excel_consolidado(
                    df_consolidado=df_consolidado,
                    parceiro_nome=parceiro_nome,
                    periodo_atual=periodo_atual,
                    dt_ini_display=dt_ini,
                    dt_fim_display=dt_fim,
                )

                if parceiro_nome == "Luggo":
                    dest_email = ["elizangela.zico@luggo.com.br"]
                elif parceiro_nome == "MRV":
                    dest_email = ["servicos@mrv.com.br"]
                elif parceiro_nome == "Selfit":
                    dest_email = [
                        "fernanda.oliveira@selfitacademias.com.br",
                        "alysson.lisboa@selfitacademias.com.br",
                        "maria.julia@selfitacademias.com.br",
                    ]
                else:
                    dest_email = [str(row["EMAIL_DESTINO"])]

                subj = f"Relatório Consolidado {parceiro_nome} - {periodo_atual}"
                html = gerar_corpo_html(parceiro_nome, periodo_atual, dt_ini, dt_fim)

                ok, msg = enviar_email(
                    dest_email,
                    subj,
                    html,
                    anexo_unico,
                    emails_copia=EMAILS_COPIA_PADRAO,
                    emails_cco=EMAILS_CCO_PADRAO,
                )
                logger.info(f"  --> Enviado para {dest_email}: {msg}")

            except Exception as e:
                logger.error(f"  Erro ao processar {parceiro_nome}: {e}")
                continue

    finally:
        try:
            conn_mysql.close()
        except Exception:
            pass

    logger.info("--- FIM DO PROCESSO v2 ---")


if __name__ == "__main__":
    main()

