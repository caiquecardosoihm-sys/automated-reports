"""
gerar_planograma.py
===================
Gera planogramas ideais por estado a partir do Snowflake.

Para cada estado:
  1. Roda a query de planograma (mesma lógica da procedure GERAR_PLANOGRAMA_ESTADO)
  2. Salva snapshot em CSV  →  snapshots/planograma_{ESTADO}_{DATA}.csv
  3. Gera Excel formatado com gráfico  →  Planograma_{ESTADO}_{DATA}.xlsx
  4. Se existir snapshot anterior, gera PDF diff  →  Diff_{ESTADO}_{DATA}.pdf

Uso:
    python gerar_planograma.py

Configuração (editar abaixo):
    ESTADOS, TOTAL_SKUS, DATA_INICIO, DATA_FIM, VOLUME_MINIMO
"""

import os
import io
import json
import logging
from datetime import date, datetime, timedelta

import snowflake.connector
from cryptography.hazmat.primitives import serialization
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
)

# ---------------------------------------------------------------------------
# Logging
# ---------------------------------------------------------------------------
logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(message)s")
logger = logging.getLogger(__name__)

SCRIPT_DIR   = os.path.dirname(os.path.abspath(__file__))
# ---------------------------------------------------------------------------
# ★ CONFIGURAÇÃO — edite aqui antes de rodar
# ---------------------------------------------------------------------------
# ESTADOS = None  → descobre automaticamente pelo Snowflake (todos os PLANOGRAMA_IDEAL_*)
# ESTADOS = ["SP", "RJ"]  → força lista específica
ESTADOS       = None
TOTAL_SKUS    = 80
DATA_FIM      = date.today()
DATA_INICIO   = DATA_FIM - timedelta(days=365)
VOLUME_MINIMO = 50   # unidades mínimas vendidas no período
# ---------------------------------------------------------------------------

SF_SCHEMA = "INHOUSE_DW.CORE_DW"
SF_CAT    = "INHOUSE_DW.ANALYTICS"

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------
def load_config():
    with open(os.path.join(SCRIPT_DIR, "config.json")) as f:
        return json.load(f)

config = load_config()

# ---------------------------------------------------------------------------
# Snowflake
# ---------------------------------------------------------------------------
def get_sf_conn():
    key_path = config["snowflake"]["private_key_file"]
    with open(key_path, "rb") as f:
        p_key = serialization.load_pem_private_key(f.read(), password=None)
    pkb = p_key.private_bytes(
        serialization.Encoding.DER,
        serialization.PrivateFormat.PKCS8,
        serialization.NoEncryption(),
    )
    return snowflake.connector.connect(
        user=config["snowflake"]["user"],
        account=config["snowflake"]["account"],
        private_key=pkb,
        warehouse=config["snowflake"]["warehouse"],
        database=config["snowflake"]["database"],
    )

def sf_query(conn, sql: str) -> pd.DataFrame:
    with conn.cursor() as cur:
        cur.execute(sql)
        rows = cur.fetchall()
        cols = [c[0] for c in cur.description]
    return pd.DataFrame(rows, columns=cols)

# ---------------------------------------------------------------------------
# Query — mesma lógica da procedure, sem CREATE TABLE
# ---------------------------------------------------------------------------
PLANOGRAMA_SQL = """
WITH
MapCategoria AS (
    SELECT CATEGORY_KEY, CATEGORY_NAME,
        CASE
            WHEN CATEGORY_NAME ILIKE '%BEBIDAS ALCOÓLICAS%' OR CATEGORY_NAME ILIKE '%CERVEJA%'
                THEN 'BEBIDAS ALCOÓLICAS'
            WHEN CATEGORY_NAME ILIKE '%BEBIDAS NÃO ALCOÓLICAS%' OR CATEGORY_NAME ILIKE '%REFRIGERANTE%'
                THEN 'BEBIDAS NÃO ALCOÓLICAS'
            ELSE CATEGORY_NAME
        END AS CATEGORIA_AGRUPADA
    FROM {SF_SCHEMA}.DIMCATEGORY
),
PerformanceAgregada AS (
    SELECT
        p.UNIFIED_PRODUCT_NAME,
        mc.CATEGORIA_AGRUPADA,
        '{estado}' AS ADDRESS_STATE,
        ANY_VALUE(p.BARCODE) AS BARCODE,
        CASE
            WHEN mc.CATEGORIA_AGRUPADA IN ('BEBIDAS ALCOÓLICAS', 'BEBIDAS NÃO ALCOÓLICAS') THEN
                CASE
                    -- Refrigerantes
                    WHEN p.UNIFIED_PRODUCT_NAME ILIKE '%COCA COLA%'   OR p.UNIFIED_PRODUCT_NAME ILIKE '%COCA-COLA%'  THEN 'MARCA_COCA_COLA'
                    WHEN p.UNIFIED_PRODUCT_NAME ILIKE '%GUARANÁ ANTARCTICA%' OR p.UNIFIED_PRODUCT_NAME ILIKE '%GUARANA ANTARCTICA%' THEN 'MARCA_GUARANA_ANTARCTICA'
                    WHEN p.UNIFIED_PRODUCT_NAME ILIKE '%PEPSI%'                                                       THEN 'MARCA_PEPSI'
                    WHEN p.UNIFIED_PRODUCT_NAME ILIKE '%FANTA%'                                                       THEN 'MARCA_FANTA'
                    WHEN p.UNIFIED_PRODUCT_NAME ILIKE '%SPRITE%'                                                      THEN 'MARCA_SPRITE'
                    WHEN p.UNIFIED_PRODUCT_NAME ILIKE '%SCHWEPPES%'                                                   THEN 'MARCA_SCHWEPPES'
                    WHEN p.UNIFIED_PRODUCT_NAME ILIKE '%H2OH%'                                                        THEN 'MARCA_H2OH'
                    WHEN p.UNIFIED_PRODUCT_NAME ILIKE '%SODA LIMONADA%' OR p.UNIFIED_PRODUCT_NAME ILIKE '%SODA ANTARC%' THEN 'MARCA_SODA_ANTARCTICA'
                    -- Cervejas
                    WHEN p.UNIFIED_PRODUCT_NAME ILIKE '%HEINEKEN%'                                                    THEN 'MARCA_HEINEKEN'
                    WHEN p.UNIFIED_PRODUCT_NAME ILIKE '%AMSTEL%'                                                      THEN 'MARCA_AMSTEL'
                    WHEN p.UNIFIED_PRODUCT_NAME ILIKE '%BRAHMA%'                                                      THEN 'MARCA_BRAHMA'
                    WHEN p.UNIFIED_PRODUCT_NAME ILIKE '%SKOL%'                                                        THEN 'MARCA_SKOL'
                    WHEN p.UNIFIED_PRODUCT_NAME ILIKE '%ANTARCTICA%'                                                  THEN 'MARCA_ANTARCTICA'
                    WHEN p.UNIFIED_PRODUCT_NAME ILIKE '%ORIGINAL%'                                                    THEN 'MARCA_ORIGINAL'
                    WHEN p.UNIFIED_PRODUCT_NAME ILIKE '%CORONA%'                                                      THEN 'MARCA_CORONA'
                    WHEN p.UNIFIED_PRODUCT_NAME ILIKE '%STELLA%'                                                      THEN 'MARCA_STELLA'
                    WHEN p.UNIFIED_PRODUCT_NAME ILIKE '%BUDWEISER%'                                                   THEN 'MARCA_BUDWEISER'
                    WHEN p.UNIFIED_PRODUCT_NAME ILIKE '%BOHEMIA%'                                                     THEN 'MARCA_BOHEMIA'
                    WHEN p.UNIFIED_PRODUCT_NAME ILIKE '%ITAIPAVA%'                                                    THEN 'MARCA_ITAIPAVA'
                    WHEN p.UNIFIED_PRODUCT_NAME ILIKE '%DEVASSA%'                                                     THEN 'MARCA_DEVASSA'
                    WHEN p.UNIFIED_PRODUCT_NAME ILIKE '%EISENBAHN%'                                                   THEN 'MARCA_EISENBAHN'
                    WHEN p.UNIFIED_PRODUCT_NAME ILIKE '%SPATEN%'                                                      THEN 'MARCA_SPATEN'
                    WHEN p.UNIFIED_PRODUCT_NAME ILIKE '%BECK%'                                                        THEN 'MARCA_BECKS'
                    WHEN p.UNIFIED_PRODUCT_NAME ILIKE '%LEFFE%'                                                       THEN 'MARCA_LEFFE'
                    WHEN p.UNIFIED_PRODUCT_NAME ILIKE '%HOEGAARDEN%'                                                  THEN 'MARCA_HOEGAARDEN'
                    ELSE 'OUTRAS_BEBIDAS'
                END
            ELSE 'OUTRAS_CATEGORIAS'
        END AS MARCA,
        SUM(s.QUANTITY)                                             AS UNIDADES_VENDIDAS_ESTADO,
        SUM(s.TOTAL_SALE_PRICE_CENTS) / 100.0                      AS FATURAMENTO_TOTAL_ESTADO,
        GREATEST(ROUND(AVG(s.QUANTITY), 0) * 7, 10)                AS QUANTIDADE_PADRAO,
        SUM(s.QUANTITY) / {dias}                                    AS VELOCIDADE_DE_VENDA_DIARIA_ESTADO,
        COUNT(DISTINCT s.STORE_DW_ID)                               AS NUM_LOJAS_DISTINTAS
    FROM {SF_SCHEMA}.FACTSALES     s
    JOIN {SF_SCHEMA}.DIMPRODUCT    p  ON s.PRODUCT_DW_ID = p.PRODUCT_DW_ID
    JOIN MapCategoria               mc ON p.CATEGORY_DW_ID = mc.CATEGORY_KEY
    JOIN {SF_SCHEMA}.DIMSTORE      st ON s.STORE_DW_ID   = st.STORE_DW_ID
    JOIN {SF_SCHEMA}.DIMDATE       d  ON s.DATE_DW_ID    = d.DATE_DW_ID
    WHERE st.ADDRESS_STATE = '{estado}'
      AND d.DATE_ACTUAL BETWEEN '{dt_ini}' AND '{dt_fim}'
    GROUP BY 1, 2, 3, 5
    HAVING SUM(s.QUANTITY) >= {vol_min}
),
ProdutosComScore AS (
    SELECT *,
        (COALESCE((FATURAMENTO_TOTAL_ESTADO - MIN(FATURAMENTO_TOTAL_ESTADO) OVER ())
            / NULLIF(MAX(FATURAMENTO_TOTAL_ESTADO) OVER () - MIN(FATURAMENTO_TOTAL_ESTADO) OVER (), 0), 0) * 0.4) +
        (COALESCE((VELOCIDADE_DE_VENDA_DIARIA_ESTADO - MIN(VELOCIDADE_DE_VENDA_DIARIA_ESTADO) OVER ())
            / NULLIF(MAX(VELOCIDADE_DE_VENDA_DIARIA_ESTADO) OVER () - MIN(VELOCIDADE_DE_VENDA_DIARIA_ESTADO) OVER (), 0), 0) * 0.4) +
        (COALESCE((NUM_LOJAS_DISTINTAS - MIN(NUM_LOJAS_DISTINTAS) OVER ())
            / NULLIF(MAX(NUM_LOJAS_DISTINTAS) OVER () - MIN(NUM_LOJAS_DISTINTAS) OVER (), 0), 0) * 0.2)
        AS PP_SCORE_ESTADO
    FROM PerformanceAgregada
),
CandidatosComVariedadeControlada AS (
    SELECT * FROM ProdutosComScore
    WHERE CATEGORIA_AGRUPADA NOT IN ('BEBIDAS ALCOÓLICAS', 'BEBIDAS NÃO ALCOÓLICAS')
    UNION ALL
    SELECT * FROM ProdutosComScore
    WHERE CATEGORIA_AGRUPADA IN ('BEBIDAS ALCOÓLICAS', 'BEBIDAS NÃO ALCOÓLICAS')
    QUALIFY ROW_NUMBER() OVER (PARTITION BY MARCA ORDER BY PP_SCORE_ESTADO DESC) <= 3
),
RankingGeral AS (
    SELECT *,
        ROW_NUMBER() OVER (ORDER BY PP_SCORE_ESTADO DESC) AS RN_GERAL
    FROM CandidatosComVariedadeControlada
)
SELECT
    ADDRESS_STATE                                                              AS ESTADO,
    CATEGORIA_AGRUPADA                                                         AS CATEGORIA,
    MARCA,
    UNIFIED_PRODUCT_NAME                                                       AS PRODUTO,
    BARCODE                                                                    AS CODIGO_BARRAS,
    ROUND(PP_SCORE_ESTADO, 4)                                                  AS SCORE,
    ROW_NUMBER() OVER (PARTITION BY CATEGORIA_AGRUPADA ORDER BY PP_SCORE_ESTADO DESC) AS RANKING_CATEGORIA,
    QUANTIDADE_PADRAO                                                           AS QTD_PADRAO,
    GREATEST(CEIL(QUANTIDADE_PADRAO * 0.3), 3)                                 AS QTD_ALERTA,
    ROUND(FATURAMENTO_TOTAL_ESTADO, 2)                                         AS VALOR_TOTAL_R,
    UNIDADES_VENDIDAS_ESTADO                                                   AS UNIDADES_VENDIDAS,
    ROUND(FATURAMENTO_TOTAL_ESTADO / NULLIF(UNIDADES_VENDIDAS_ESTADO, 0), 2)  AS PRECO_MEDIO_R,
    ROUND(VELOCIDADE_DE_VENDA_DIARIA_ESTADO, 2)                               AS VENDA_DIARIA,
    NUM_LOJAS_DISTINTAS                                                        AS LOJAS
FROM RankingGeral
WHERE RN_GERAL <= {total_skus}
ORDER BY CATEGORIA, RANKING_CATEGORIA
"""

def rodar_planograma(conn, estado: str, total_skus: int = TOTAL_SKUS) -> pd.DataFrame:
    dias = (DATA_FIM - DATA_INICIO).days + 1
    sql = PLANOGRAMA_SQL.format(
        SF_SCHEMA=SF_SCHEMA,
        estado=estado,
        dias=dias,
        dt_ini=DATA_INICIO,
        dt_fim=DATA_FIM,
        vol_min=VOLUME_MINIMO,
        total_skus=total_skus,
    )
    logger.info(f"  [{estado}] Rodando query Snowflake ...")
    df = sf_query(conn, sql)
    logger.info(f"  [{estado}] {len(df)} SKUs retornados")
    return df

# ---------------------------------------------------------------------------
# Snapshot — lê/escreve em INHOUSE_DW.ANALYTICS.PLANOGRAMA_IDEAL_{ESTADO}
# ---------------------------------------------------------------------------
SF_ANALYTICS = "INHOUSE_DW.ANALYTICS"

COL_NORMALIZE = {
    "SCORE_LOCAL":           "SCORE",
    "QUANTIDADE_PADRAO":     "QTD_PADRAO",
    "QUANTIDADE_ALERTA":     "QTD_ALERTA",
    "VALOR_TOTAL_VENDIDO_R": "VALOR_TOTAL_R",
    "QTD_TOTAL_ITENS_VENDIDOS": "UNIDADES_VENDIDAS",
    "RANKING_NA_CATEGORIA":  "RANKING_CATEGORIA",
}

def carregar_snapshot_anterior(conn, estado: str) -> pd.DataFrame | None:
    tabela = f"{SF_ANALYTICS}.PLANOGRAMA_IDEAL_{estado}"
    try:
        df = sf_query(conn, f"SELECT * FROM {tabela}")
        if df.empty:
            return None
        df = df.rename(columns=COL_NORMALIZE)
        logger.info(f"  [{estado}] Snapshot anterior carregado → {tabela} ({len(df)} SKUs)")
        return df
    except Exception:
        logger.info(f"  [{estado}] Sem snapshot anterior no Snowflake.")
        return None

def salvar_snapshot(conn, df: pd.DataFrame, estado: str):
    """Sobrescreve PLANOGRAMA_IDEAL_{ESTADO} com o planograma novo."""
    tabela = f"{SF_ANALYTICS}.PLANOGRAMA_IDEAL_{estado}"
    # Cria tabela temporária e faz INSERT via write_pandas
    from snowflake.connector.pandas_tools import write_pandas
    df_sf = df.copy()
    df_sf.columns = [c.upper() for c in df_sf.columns]
    with conn.cursor() as cur:
        cur.execute(f"CREATE OR REPLACE TABLE {tabela} AS SELECT * FROM VALUES (1) v(x) WHERE 1=0")
        # Recria estrutura correta
        cols_ddl = ", ".join(f'"{c}" VARCHAR' for c in df_sf.columns)
        cur.execute(f"CREATE OR REPLACE TABLE {tabela} ({cols_ddl})")
    success, nchunks, nrows, _ = write_pandas(conn, df_sf,
                                               table_name=f"PLANOGRAMA_IDEAL_{estado}",
                                               database="INHOUSE_DW",
                                               schema="ANALYTICS",
                                               overwrite=True,
                                               auto_create_table=True)
    logger.info(f"  [{estado}] Snapshot salvo → {tabela} ({nrows} linhas)")

# ---------------------------------------------------------------------------
# Excel do planograma
# ---------------------------------------------------------------------------
COR_HEADER  = "1A1A2E"
COR_ALT     = "F0F4FF"
COR_VERDE   = "D6F0D6"
COR_VERMELHO= "FFD6D6"

BORDER = Border(
    left=Side(style="thin", color="CCCCCC"), right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),  bottom=Side(style="thin", color="CCCCCC"),
)

def _estilo_header(cell):
    cell.font      = Font(bold=True, color="FFFFFF", size=10, name="Arial")
    cell.fill      = PatternFill("solid", fgColor=COR_HEADER)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = BORDER

def _largura_col(df, col_idx):
    col_name = df.columns[col_idx]
    max_len  = max(len(str(col_name)),
                   df.iloc[:, col_idx].astype(str).str.len().max() if not df.empty else 0)
    return min(max_len + 3, 45)

def gerar_excel_planograma(df: pd.DataFrame, estado: str, hoje: date) -> str:
    path = os.path.join(SCRIPT_DIR, f"Planograma_{estado}_{hoje.strftime('%Y%m%d')}.xlsx")

    # Garante tipos numéricos nas colunas calculadas
    for col in ["VALOR_TOTAL_R", "UNIDADES_VENDIDAS", "SCORE", "QTD_PADRAO",
                "QTD_ALERTA", "PRECO_MEDIO_R", "VENDA_DIARIA", "LOJAS"]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")

    # ── Aba 1: Planograma completo ──────────────────────────────────────────
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Planograma", index=False)

        # Aba 2: Resumo por categoria
        resumo = (df.groupby("CATEGORIA", as_index=False)
                  .agg(SKUs=("PRODUTO","count"),
                       Unidades=("UNIDADES_VENDIDAS","sum"),
                       Faturamento=("VALOR_TOTAL_R","sum")))
        resumo["% SKUs"]    = (resumo["SKUs"] / resumo["SKUs"].sum() * 100).round(1)
        resumo["% Faturamento"] = (resumo["Faturamento"] / resumo["Faturamento"].sum() * 100).round(1)
        resumo = resumo.sort_values("Faturamento", ascending=False)
        resumo.to_excel(writer, sheet_name="Resumo por Categoria", index=False)

    buf.seek(0)
    wb = load_workbook(buf)

    # ── Formatar aba Planograma ─────────────────────────────────────────────
    ws = wb["Planograma"]
    ws.freeze_panes = "A2"
    for col_idx, col_name in enumerate(df.columns, 1):
        _estilo_header(ws.cell(row=1, column=col_idx))
        ws.column_dimensions[get_column_letter(col_idx)].width = _largura_col(df, col_idx - 1)

    for row_idx, row in enumerate(df.itertuples(index=False), 2):
        fill = PatternFill("solid", fgColor=COR_ALT) if row_idx % 2 == 0 else PatternFill()
        for col_idx, val in enumerate(row, 1):
            cell        = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = BORDER
            cell.fill   = fill
            cell.font   = Font(name="Arial", size=9)
            col_name = df.columns[col_idx - 1]
            if col_name in ("VALOR_TOTAL_R", "PRECO_MEDIO_R"):
                cell.number_format = "#,##0.00"
                cell.alignment = Alignment(horizontal="right")
            elif col_name in ("UNIDADES_VENDIDAS", "QTD_PADRAO", "QTD_ALERTA", "LOJAS"):
                cell.number_format = "#,##0"
                cell.alignment = Alignment(horizontal="right")
            elif col_name == "SCORE":
                cell.number_format = "0.0000"
                cell.alignment = Alignment(horizontal="right")

    ws.row_dimensions[1].height = 32

    # ── Formatar + Gráfico aba Resumo ───────────────────────────────────────
    ws2 = wb["Resumo por Categoria"]
    ws2.freeze_panes = "A2"
    for col_idx, col_name in enumerate(resumo.columns, 1):
        _estilo_header(ws2.cell(row=1, column=col_idx))
        ws2.column_dimensions[get_column_letter(col_idx)].width = _largura_col(resumo, col_idx - 1)

    for row_idx, row in enumerate(resumo.itertuples(index=False), 2):
        fill = PatternFill("solid", fgColor=COR_ALT) if row_idx % 2 == 0 else PatternFill()
        for col_idx, val in enumerate(row, 1):
            cell        = ws2.cell(row=row_idx, column=col_idx, value=val)
            cell.border = BORDER
            cell.fill   = fill
            cell.font   = Font(name="Arial", size=9)
            col_name = resumo.columns[col_idx - 1]
            if col_name == "Faturamento":
                cell.number_format = "#,##0.00"
                cell.alignment = Alignment(horizontal="right")
            elif col_name in ("SKUs", "Unidades"):
                cell.number_format = "#,##0"
                cell.alignment = Alignment(horizontal="right")
            elif "%" in col_name:
                cell.number_format = "0.0"
                cell.alignment = Alignment(horizontal="right")

    # Gráfico — Faturamento por categoria
    n_rows = len(resumo) + 1
    chart  = BarChart()
    chart.type    = "bar"
    chart.title   = f"Faturamento por Categoria — {estado}"
    chart.y_axis.title = "R$"
    chart.shape   = 4
    chart.width   = 20
    chart.height  = 14

    data_ref = Reference(ws2, min_col=4, min_row=1, max_row=n_rows)   # col Faturamento = 4 (A=cat,B=SKUs,C=Unidades,D=Fat)
    cats_ref = Reference(ws2, min_col=1, min_row=2, max_row=n_rows)

    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.series[0].graphicalProperties.solidFill = "1A1A2E"

    ws2.add_chart(chart, f"A{n_rows + 3}")

    wb.save(path)
    logger.info(f"  [{estado}] Excel → {os.path.basename(path)}")
    return path

# ---------------------------------------------------------------------------
# PDF Diff
# ---------------------------------------------------------------------------
def gerar_pdf_diff(df_novo: pd.DataFrame, df_ant: pd.DataFrame,
                   estado: str, hoje: date, data_ant: str) -> str:

    path = os.path.join(SCRIPT_DIR, f"Diff_Planograma_{estado}_{hoje.strftime('%Y%m%d')}.pdf")

    skus_novo = set(df_novo["PRODUTO"].str.strip().str.upper())
    skus_ant  = set(df_ant["PRODUTO"].str.strip().str.upper())

    entrou  = skus_novo - skus_ant
    saiu    = skus_ant  - skus_novo
    manteve = skus_novo & skus_ant

    df_entrou = df_novo[df_novo["PRODUTO"].str.strip().str.upper().isin(entrou)].copy()
    df_saiu   = df_ant [df_ant ["PRODUTO"].str.strip().str.upper().isin(saiu)  ].copy()

    # ReportLab
    doc    = SimpleDocTemplate(path, pagesize=A4,
                                topMargin=2*cm, bottomMargin=2*cm,
                                leftMargin=2*cm, rightMargin=2*cm)
    styles = getSampleStyleSheet()
    story  = []

    # Estilo customizado
    titulo_st = ParagraphStyle("titulo", parent=styles["Title"],
                                fontSize=18, textColor=colors.HexColor("#1A1A2E"),
                                spaceAfter=6)
    sub_st    = ParagraphStyle("sub", parent=styles["Normal"],
                                fontSize=10, textColor=colors.grey, spaceAfter=16)
    h2_st     = ParagraphStyle("h2", parent=styles["Heading2"],
                                fontSize=13, textColor=colors.HexColor("#1A1A2E"),
                                spaceBefore=18, spaceAfter=8)
    normal_st = styles["Normal"]
    normal_st.fontName = "Helvetica"

    # Cabeçalho
    story.append(Paragraph(f"📋 Diff de Planograma — {estado}", titulo_st))
    story.append(Paragraph(
        f"Comparação: <b>{data_ant}</b>  →  <b>{hoje.strftime('%d/%m/%Y')}</b>  |  "
        f"Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')}",
        sub_st))
    story.append(HRFlowable(width="100%", thickness=2, color=colors.HexColor("#1A1A2E")))
    story.append(Spacer(1, 0.4*cm))

    # Resumo executivo
    story.append(Paragraph("Resumo", h2_st))
    resumo_data = [
        ["", "Qtd", ""],
        ["✅  Entraram no planograma", str(len(entrou)),  ""],
        ["❌  Saíram do planograma",   str(len(saiu)),    ""],
        ["⟳   Permaneceram",           str(len(manteve)), ""],
        ["Total de SKUs (novo)",        str(len(skus_novo)), ""],
    ]
    ts = TableStyle([
        ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#1A1A2E")),
        ("TEXTCOLOR",  (0,0), (-1,0), colors.white),
        ("FONTNAME",   (0,0), (-1,0), "Helvetica-Bold"),
        ("FONTSIZE",   (0,0), (-1,-1), 10),
        ("ROWBACKGROUNDS", (0,1), (-1,-1), [colors.HexColor("#F0F4FF"), colors.white]),
        ("GRID",       (0,0), (-1,-1), 0.5, colors.HexColor("#CCCCCC")),
        ("ALIGN",      (1,0), (1,-1), "CENTER"),
        ("FONTNAME",   (0,1), (-1,-1), "Helvetica"),
        ("TOPPADDING",  (0,0), (-1,-1), 6),
        ("BOTTOMPADDING",(0,0),(-1,-1), 6),
    ])
    t = Table(resumo_data, colWidths=[11*cm, 3*cm, 3*cm])
    t.setStyle(ts)
    story.append(t)
    story.append(Spacer(1, 0.6*cm))

    # Produtos que ENTRARAM
    if not df_entrou.empty:
        story.append(Paragraph(f"✅  Entraram ({len(df_entrou)} SKUs)", h2_st))
        cols_show = ["PRODUTO", "CATEGORIA", "SCORE", "QTD_PADRAO", "PRECO_MEDIO_R"]
        df_e = df_entrou[cols_show].sort_values("SCORE", ascending=False)
        header = [["Produto", "Categoria", "Score", "Qtd Padrão", "Preço Médio R$"]]
        rows   = [[str(v) for v in row] for row in df_e.itertuples(index=False)]
        tbl_data = header + rows
        ts_e = TableStyle([
            ("BACKGROUND",   (0,0), (-1,0), colors.HexColor("#1C6B3A")),
            ("TEXTCOLOR",    (0,0), (-1,0), colors.white),
            ("FONTNAME",     (0,0), (-1,0), "Helvetica-Bold"),
            ("FONTSIZE",     (0,0), (-1,-1), 8),
            ("ROWBACKGROUNDS",(0,1),(-1,-1), [colors.HexColor("#D6F0D6"), colors.white]),
            ("GRID",         (0,0), (-1,-1), 0.5, colors.HexColor("#CCCCCC")),
            ("FONTNAME",     (0,1), (-1,-1), "Helvetica"),
            ("TOPPADDING",   (0,0), (-1,-1), 4),
            ("BOTTOMPADDING",(0,0), (-1,-1), 4),
            ("WORDWRAP",     (0,0), (-1,-1), True),
        ])
        t_e = Table(tbl_data, colWidths=[7*cm, 4*cm, 2*cm, 2.5*cm, 2.5*cm])
        t_e.setStyle(ts_e)
        story.append(t_e)
        story.append(Spacer(1, 0.4*cm))

    # Produtos que SAÍRAM
    if not df_saiu.empty:
        story.append(Paragraph(f"❌  Saíram ({len(df_saiu)} SKUs)", h2_st))
        cols_show = ["PRODUTO", "CATEGORIA", "SCORE", "QTD_PADRAO", "PRECO_MEDIO_R"]
        df_s = df_saiu[cols_show].sort_values("SCORE", ascending=False)
        header = [["Produto", "Categoria", "Score", "Qtd Padrão", "Preço Médio R$"]]
        rows   = [[str(v) for v in row] for row in df_s.itertuples(index=False)]
        tbl_data = header + rows
        ts_s = TableStyle([
            ("BACKGROUND",   (0,0), (-1,0), colors.HexColor("#8B1C1C")),
            ("TEXTCOLOR",    (0,0), (-1,0), colors.white),
            ("FONTNAME",     (0,0), (-1,0), "Helvetica-Bold"),
            ("FONTSIZE",     (0,0), (-1,-1), 8),
            ("ROWBACKGROUNDS",(0,1),(-1,-1), [colors.HexColor("#FFD6D6"), colors.white]),
            ("GRID",         (0,0), (-1,-1), 0.5, colors.HexColor("#CCCCCC")),
            ("FONTNAME",     (0,1), (-1,-1), "Helvetica"),
            ("TOPPADDING",   (0,0), (-1,-1), 4),
            ("BOTTOMPADDING",(0,0), (-1,-1), 4),
            ("WORDWRAP",     (0,0), (-1,-1), True),
        ])
        t_s = Table(tbl_data, colWidths=[7*cm, 4*cm, 2*cm, 2.5*cm, 2.5*cm])
        t_s.setStyle(ts_s)
        story.append(t_s)

    doc.build(story)
    logger.info(f"  [{estado}] PDF diff → {os.path.basename(path)}")
    return path


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    logger.info("=" * 55)
    logger.info("  InHouse Market — Gerador de Planogramas")
    logger.info(f"  Período: {DATA_INICIO} → {DATA_FIM}")
    logger.info(f"  Estados: {', '.join(ESTADOS) if ESTADOS else 'todos (auto-descoberta)'}")
    logger.info(f"  SKUs por estado: {TOTAL_SKUS}  |  Vol. mín.: {VOLUME_MINIMO}")
    logger.info("=" * 55)

    hoje = date.today()
    conn = get_sf_conn()

    try:
        # Descobre estados automaticamente se não estiver hardcoded
        estados = ESTADOS
        if estados is None:
            df_tabs = sf_query(conn, """
                SELECT REPLACE(TABLE_NAME, 'PLANOGRAMA_IDEAL_', '') AS ESTADO
                FROM INHOUSE_DW.INFORMATION_SCHEMA.TABLES
                WHERE TABLE_SCHEMA = 'ANALYTICS'
                  AND TABLE_NAME LIKE 'PLANOGRAMA_IDEAL_%'
                ORDER BY ESTADO
            """)
            estados = df_tabs["ESTADO"].tolist()
            logger.info(f"  Estados encontrados: {', '.join(estados)}")

        for estado in estados:
            logger.info(f"\n[ {estado} ]")

            # 1. Carrega snapshot anterior para saber o total de SKUs
            df_ant   = carregar_snapshot_anterior(conn, estado)
            n_skus   = len(df_ant) if df_ant is not None else TOTAL_SKUS

            # 3. Roda query com o mesmo número de SKUs do snapshot anterior
            df = rodar_planograma(conn, estado, n_skus)
            if df.empty:
                logger.warning(f"  [{estado}] Sem dados — pulando.")
                continue

            # 4. Salva novo planograma no Snowflake (sobrescreve)
            salvar_snapshot(conn, df, estado)

            # 5. Excel
            gerar_excel_planograma(df, estado, hoje)

            # 6. Diff PDF (só se tiver snapshot anterior)
            if df_ant is not None:
                try:
                    gerar_pdf_diff(df, df_ant, estado, hoje, "snapshot anterior")
                except PermissionError:
                    logger.warning(f"  [{estado}] PDF bloqueado (arquivo aberto?) — feche e rode só esse estado depois.")
            else:
                logger.info(f"  [{estado}] Sem snapshot anterior — diff será gerado na próxima rodada.")

    finally:
        conn.close()

    logger.info("\n" + "=" * 55)
    logger.info("  Concluído.")
    logger.info("=" * 55)


if __name__ == "__main__":
    main()
