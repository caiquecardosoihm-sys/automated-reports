"""
data_pull_pascoa_e_energeticos.py
==================================
Gera dois data pulls do Snowflake (INHOUSE_DW.CORE_DW) e exporta para Excel.

REPORT 1 — Chocolates Páscoa (2 abas)
  Aba A: Faturamento chocolates 02/04/2026 a 05/04/2026
          vs 26/03/2026 a 29/03/2026 (semana anterior à Páscoa)
          Apenas lojas com vendas nos dois períodos.

  Aba B: Faturamento chocolates 02/04/2026 a 05/04/2026
          vs 02/04/2025 a 05/04/2025 (mesmo período ano anterior)
          Apenas lojas com vendas nos dois anos.

REPORT 2 — Energéticos últimos 12 meses
  Por mês, Estado, Cidade, SKU, Produto, Quantidade, Valor.
  Filtro por nome via ILIKE (Monster, Red Bull, Baly, etc.).

Uso:
    python data_pull_pascoa_e_energeticos.py

Saída:
    Chocolates_Pascoa_YYYYMMDD.xlsx
    Energeticos_Anual_YYYYMMDD.xlsx
    (na mesma pasta do script)
"""

import os
import json
import io
import logging
from datetime import date, timedelta

import snowflake.connector
from cryptography.hazmat.primitives import serialization
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Logging
# ---------------------------------------------------------------------------
logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(message)s")
logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Config — lê config.json na mesma pasta do script
# ---------------------------------------------------------------------------
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))

def load_config(path: str = None) -> dict:
    path = path or os.path.join(SCRIPT_DIR, "config.json")
    with open(path, "r") as f:
        return json.load(f)

config = load_config()

# Schema Snowflake
SF_SCHEMA = "INHOUSE_DW.CORE_DW"

# ---------------------------------------------------------------------------
# Conexão Snowflake via RSA key
# ---------------------------------------------------------------------------
def get_snowflake_conn():
    key_path = config["snowflake"]["private_key_file"]
    with open(key_path, "rb") as key_file:
        p_key = serialization.load_pem_private_key(key_file.read(), password=None)
    pkb = p_key.private_bytes(
        encoding=serialization.Encoding.DER,
        format=serialization.PrivateFormat.PKCS8,
        encryption_algorithm=serialization.NoEncryption(),
    )
    conn = snowflake.connector.connect(
        user=config["snowflake"]["user"],
        account=config["snowflake"]["account"],
        private_key=pkb,
        warehouse=config["snowflake"]["warehouse"],
        database=config["snowflake"]["database"],
    )
    logger.info("Conexão com Snowflake estabelecida.")
    return conn


def sf_query(conn, sql: str) -> pd.DataFrame:
    """Executa uma query no Snowflake e retorna DataFrame."""
    with conn.cursor() as cur:
        cur.execute(sql)
        rows = cur.fetchall()
        cols = [c[0] for c in cur.description]
    return pd.DataFrame(rows, columns=cols)


# ---------------------------------------------------------------------------
# Filtro de chocolates via CATEGORY_DW_ID (diagnóstico confirmou as chaves)
# Muito mais confiável que ILIKE — captura "KIT-KAT TRADICIONAL", "SNICKERS",
# "BOMBOM GAROTO" etc. independente de como o produto foi nomeado.
#
# Categorias incluídas (mapeadas em DIMCATEGORY):
#   620  CHOCOLATES              | 664  DOCES/CHOCOLATES
#   137  CHOCOLATES / DOCES      | 1334 CHOCOLATE
#   1357 DOCES E CHOCOLATES      | 68378 CHOCOLATES/BALAS
#   1464 BOMBONIERE - CHOCOLATE  | 605  CHOCOLATES E DOCES
#   1173 BARRAS CHOCOLATE/BOMBOM | 312  CACAU SHOW CHOCOLATES
#   68430 CACAU SHOW CHOCOLATERIA
#
# Excluídas intencionalmente: achocolatados líquidos/em pó (bebidas).
# ---------------------------------------------------------------------------
CHOCOLATE_CATEGORY_KEYS = (
    620, 664, 137, 1334, 1357, 68378, 1464, 605, 1173, 312, 68430
)

CHOC_CATEGORY_FILTER = (
    f"p.CATEGORY_DW_ID IN ({', '.join(str(k) for k in CHOCOLATE_CATEGORY_KEYS)})"
)

# ---------------------------------------------------------------------------
# Padrões ILIKE para energéticos (ILIKE ainda é adequado aqui pois não há
# categoria confiável equivalente na DIMCATEGORY)
# ---------------------------------------------------------------------------
ENERGETICO_ILIKES = [
    "%monster%", "%red bull%", "%redbull%", "%baly%",
    "%burn%", "%hell energy%", "%tnt energy%", "%flash power%",
    "%bang energy%", "%reign%", "%rockstar%", "%battery%",
    "%dark dog%", "%220v%", "%energetico%", "%energético%",
    "%energy drink%",
]

def build_ilike_filter(column: str, patterns: list) -> str:
    """Monta bloco OR de ILIKE para uma lista de padrões."""
    clauses = "\n             OR ".join(f"{column} ILIKE '{p}'" for p in patterns)
    return f"({clauses})"


# ---------------------------------------------------------------------------
# REPORT 1 — Chocolates Páscoa
# ---------------------------------------------------------------------------

def query_chocolates_sf(conn, dt_ini: date, dt_fim: date, lojas_cte: str) -> pd.DataFrame:
    """
    Puxa faturamento de chocolates por LOJA no período informado,
    limitando às lojas da CTE 'lojas_ambos'.
    Agrega no nível de loja — sem descer para produto/SKU.
    Filtro por CATEGORY_DW_ID (muito mais preciso que ILIKE por nome).
    """
    sql = f"""
    {lojas_cte}
    SELECT
        st.STORE_NAME                                    AS "Loja",
        st.ADDRESS_STATE                                 AS "Estado",
        st.ADDRESS_CITY                                  AS "Cidade",
        SUM(f.QUANTITY)                                  AS "Quantidade",
        ROUND(SUM(f.TOTAL_SALE_PRICE_CENTS) / 100.0, 2) AS "Valor_R$"
    FROM {SF_SCHEMA}.FACTSALES        f
    JOIN {SF_SCHEMA}.DIMDATE          d  ON d.DATE_DW_ID    = f.DATE_DW_ID
    JOIN {SF_SCHEMA}.DIMSTORE         st ON st.STORE_DW_ID  = f.STORE_DW_ID
    JOIN {SF_SCHEMA}.DIMPRODUCT       p  ON p.PRODUCT_DW_ID = f.PRODUCT_DW_ID
    WHERE f.STORE_DW_ID IN (SELECT STORE_DW_ID FROM lojas_ambos)
      AND d.DATE_ACTUAL BETWEEN '{dt_ini}' AND '{dt_fim}'
      AND {CHOC_CATEGORY_FILTER}
    GROUP BY st.STORE_NAME, st.ADDRESS_STATE, st.ADDRESS_CITY
    ORDER BY "Estado", "Cidade", "Loja"
    """
    return sf_query(conn, sql)


def build_lojas_ambos_cte(per_a_ini: date, per_a_fim: date,
                           per_b_ini: date, per_b_fim: date,
                           ilike_filter: str = None) -> str:
    """
    CTE que retorna STORE_DW_ID das lojas com pelo menos 1 venda
    em AMBOS os períodos. Se ilike_filter for passado, restringe
    às lojas que venderam aquele grupo de produtos nos dois períodos.
    """
    produto_join  = f"JOIN {SF_SCHEMA}.DIMPRODUCT p ON p.PRODUCT_DW_ID = f.PRODUCT_DW_ID"
    produto_where = f"AND {ilike_filter}" if ilike_filter else ""

    return f"""
    WITH lojas_per_a AS (
        SELECT DISTINCT f.STORE_DW_ID
        FROM {SF_SCHEMA}.FACTSALES f
        JOIN {SF_SCHEMA}.DIMDATE   d ON d.DATE_DW_ID = f.DATE_DW_ID
        {produto_join}
        WHERE d.DATE_ACTUAL BETWEEN '{per_a_ini}' AND '{per_a_fim}'
        {produto_where}
    ),
    lojas_per_b AS (
        SELECT DISTINCT f.STORE_DW_ID
        FROM {SF_SCHEMA}.FACTSALES f
        JOIN {SF_SCHEMA}.DIMDATE   d ON d.DATE_DW_ID = f.DATE_DW_ID
        {produto_join}
        WHERE d.DATE_ACTUAL BETWEEN '{per_b_ini}' AND '{per_b_fim}'
        {produto_where}
    ),
    lojas_ambos AS (
        SELECT STORE_DW_ID FROM lojas_per_a
        INTERSECT
        SELECT STORE_DW_ID FROM lojas_per_b
    )"""


def pull_chocolates_pascoa(conn) -> dict:
    logger.info("=== REPORT 1: Chocolates Páscoa ===")
    results = {}

    # --- Aba A: Páscoa 2026 vs semana anterior ---
    p26_ini, p26_fim   = date(2026, 4, 2), date(2026, 4, 5)
    sem_ini, sem_fim   = date(2026, 3, 26), date(2026, 3, 29)

    logger.info(f"  Aba A: {p26_ini} a {p26_fim}  vs  {sem_ini} a {sem_fim}")
    cte_a = build_lojas_ambos_cte(p26_ini, p26_fim, sem_ini, sem_fim, ilike_filter=CHOC_CATEGORY_FILTER)

    df_pascoa_a = query_chocolates_sf(conn, p26_ini, p26_fim, cte_a)
    df_semana_a = query_chocolates_sf(conn, sem_ini,  sem_fim,  cte_a)

    df_pascoa_a.insert(0, "Periodo", f"{p26_ini.strftime('%d/%m')} a {p26_fim.strftime('%d/%m/%Y')}  — Páscoa 2026")
    df_semana_a.insert(0, "Periodo", f"{sem_ini.strftime('%d/%m')} a {sem_fim.strftime('%d/%m/%Y')}  — Semana anterior")

    df_aba_a = pd.concat([df_pascoa_a, df_semana_a], ignore_index=True)
    logger.info(f"  Aba A: {len(df_pascoa_a)} linhas Páscoa + {len(df_semana_a)} linhas semana anterior")
    results["Pascoa_vs_Semana_Anterior"] = df_aba_a

    # --- Aba B: Páscoa 2026 vs Páscoa 2025 ---
    p25_ini, p25_fim   = date(2025, 4, 2), date(2025, 4, 5)

    logger.info(f"  Aba B: {p26_ini} a {p26_fim}  vs  {p25_ini} a {p25_fim}")
    cte_b = build_lojas_ambos_cte(p26_ini, p26_fim, p25_ini, p25_fim, ilike_filter=CHOC_CATEGORY_FILTER)

    df_2026 = query_chocolates_sf(conn, p26_ini, p26_fim, cte_b)
    df_2025 = query_chocolates_sf(conn, p25_ini, p25_fim, cte_b)

    df_2026.insert(0, "Periodo", f"{p26_ini.strftime('%d/%m')} a {p26_fim.strftime('%d/%m/%Y')}  — Páscoa 2026")
    df_2025.insert(0, "Periodo", f"{p25_ini.strftime('%d/%m')} a {p25_fim.strftime('%d/%m/%Y')}  — Páscoa 2025")

    df_aba_b = pd.concat([df_2026, df_2025], ignore_index=True)
    logger.info(f"  Aba B: {len(df_2026)} linhas 2026 + {len(df_2025)} linhas 2025")
    results["Pascoa_2026_vs_2025"] = df_aba_b

    return results


# ---------------------------------------------------------------------------
# REPORT 2 — Energéticos últimos 12 meses
# ---------------------------------------------------------------------------

def pull_energeticos_anual(conn) -> pd.DataFrame:
    logger.info("=== REPORT 2: Energéticos Anuais ===")

    hoje   = date.today()
    dt_fim = hoje - timedelta(days=1)
    dt_ini = date(dt_fim.year - 1, dt_fim.month, dt_fim.day)

    logger.info(f"  Período: {dt_ini} a {dt_fim}")

    ilike_filter = build_ilike_filter("p.UNIFIED_PRODUCT_NAME", ENERGETICO_ILIKES)

    sql = f"""
    SELECT
        TO_CHAR(d.DATE_ACTUAL, 'YYYY-MM')               AS "Mes",
        st.ADDRESS_STATE                                 AS "Estado",
        st.ADDRESS_CITY                                  AS "Cidade",
        p.PRODUCT_OLTP_ID                                AS "SKU",
        p.UNIFIED_PRODUCT_NAME                           AS "Produto",
        SUM(f.QUANTITY)                                  AS "Quantidade",
        ROUND(SUM(f.TOTAL_SALE_PRICE_CENTS) / 100.0, 2) AS "Valor_R$"
    FROM {SF_SCHEMA}.FACTSALES        f
    JOIN {SF_SCHEMA}.DIMDATE          d  ON d.DATE_DW_ID    = f.DATE_DW_ID
    JOIN {SF_SCHEMA}.DIMSTORE         st ON st.STORE_DW_ID  = f.STORE_DW_ID
    JOIN {SF_SCHEMA}.DIMPRODUCT       p  ON p.PRODUCT_DW_ID = f.PRODUCT_DW_ID
    WHERE d.DATE_ACTUAL BETWEEN '{dt_ini}' AND '{dt_fim}'
      AND {ilike_filter}
    GROUP BY "Mes", st.ADDRESS_STATE, st.ADDRESS_CITY,
             p.PRODUCT_OLTP_ID, p.UNIFIED_PRODUCT_NAME
    ORDER BY "Mes", "Estado", "Cidade", "Valor_R$" DESC
    """

    df = sf_query(conn, sql)
    logger.info(f"  Energéticos: {len(df)} linhas encontradas")
    return df


# ---------------------------------------------------------------------------
# Formatação Excel
# ---------------------------------------------------------------------------
HEADER_FILL = PatternFill("solid", fgColor="1A1A2E")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
ALT_FILL    = PatternFill("solid", fgColor="F5F5F5")
BORDER      = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)


def format_sheet(ws, df: pd.DataFrame):
    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value     = col_name
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = BORDER

    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        fill = ALT_FILL if row_idx % 2 == 0 else PatternFill()
        for col_idx, value in enumerate(row, start=1):
            cell       = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = BORDER
            cell.fill   = fill
            col_name    = df.columns[col_idx - 1]
            if col_name in ("Valor_R$",):
                cell.number_format = "#,##0.00"
                cell.alignment = Alignment(horizontal="right")
            elif col_name in ("Quantidade", "SKU"):
                cell.number_format = "#,##0"
                cell.alignment = Alignment(horizontal="right")

    for col_idx in range(1, len(df.columns) + 1):
        col_name = df.columns[col_idx - 1]
        max_len  = max(
            len(str(col_name)),
            df.iloc[:, col_idx - 1].astype(str).str.len().max() if not df.empty else 0,
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 55)

    ws.freeze_panes = "A2"


def salvar_excel(sheets: dict, filepath: str):
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        for name, df in sheets.items():
            df.to_excel(writer, sheet_name=name[:31], index=False)

    buffer.seek(0)
    wb = load_workbook(buffer)
    for name, df in sheets.items():
        format_sheet(wb[name[:31]], df)

    wb.save(filepath)
    logger.info(f"  Salvo: {filepath}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    logger.info("=" * 50)
    logger.info("  InHouse Market — Data Pull (Snowflake)")
    logger.info("=" * 50)

    hoje = date.today()
    conn = get_snowflake_conn()

    try:
        # Report 1: Chocolates Páscoa
        chocolates = pull_chocolates_pascoa(conn)
        arq_choc = os.path.join(SCRIPT_DIR, f"Chocolates_Pascoa_{hoje.strftime('%Y%m%d')}.xlsx")
        salvar_excel(chocolates, arq_choc)

        # Report 2: Energéticos Anuais
        df_energ = pull_energeticos_anual(conn)
        arq_energ = os.path.join(SCRIPT_DIR, f"Energeticos_Anual_{hoje.strftime('%Y%m%d')}.xlsx")
        salvar_excel({"Energeticos_12meses": df_energ}, arq_energ)

    finally:
        conn.close()

    logger.info("=" * 50)
    logger.info("  Concluído. Arquivos gerados:")
    logger.info(f"    {os.path.basename(arq_choc)}")
    logger.info(f"    {os.path.basename(arq_energ)}")
    logger.info("=" * 50)


if __name__ == "__main__":
    main()
