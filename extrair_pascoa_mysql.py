"""
extrair_pascoa_mysql.py
=======================
Estratégia ETL: sem JOIN no MySQL, puxa cada tabela separada e junta no pandas.
Assim evita o max_statement_execution_time do servidor.

Fluxo:
1. Puxa sales_items / sales por período (sem JOIN)
2. Puxa stores e addresses uma vez (tabelas pequenas)
3. Tudo salvo em .parquet como cache local
4. Pandas faz o merge, filtro de chocolate e agregação
5. Gera Excel formatado com duas abas

Uso:
    python extrair_pascoa_mysql.py
"""

import os
import io
import json
import logging
from datetime import date

import mysql.connector
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Logging
# ---------------------------------------------------------------------------
logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(message)s")
logger = logging.getLogger(__name__)

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------
def load_config():
    path = os.path.join(SCRIPT_DIR, "config.json")
    with open(path) as f:
        return json.load(f)

config = load_config()

# ---------------------------------------------------------------------------
# Períodos
# ---------------------------------------------------------------------------
PERIODOS = {
    "pascoa_2026": ("2026-04-02", "2026-04-05"),
    "semana_ant":  ("2026-03-26", "2026-03-29"),
    "pascoa_2025": ("2025-04-02", "2025-04-05"),
}

# ---------------------------------------------------------------------------
# Filtro de chocolate por category_name (direto no sales_items)
# ---------------------------------------------------------------------------
CHOC_CATEGORIES = [
    "%chocolate%", "%choco%", "%bomboniere%",
    "%bombom%",    "%cacau%", "%trufas%",
    "%snickers%",  "%kit-kat%", "%kit kat%", "%kitkat%",
    "%ferrero%",   "%lacta%",   "%garoto%",
]

CHOC_FILTER = " OR ".join(f"si.category_name LIKE '{p}'" for p in CHOC_CATEGORIES)

# ---------------------------------------------------------------------------
# Conexão MySQL — igual ao ETL de referência
# ---------------------------------------------------------------------------
def get_mysql_conn():
    m = config["mysql"]
    return mysql.connector.connect(
        host=m["host"],
        user=m["user"],
        password=m["password"],
        database=m["database"],
        connection_timeout=30,
    )

# ---------------------------------------------------------------------------
# Extração sem JOIN — mesma estratégia do ETL (pd.read_sql por tabela)
# ---------------------------------------------------------------------------
def extrair_tabela(conn, nome_parquet: str, sql: str) -> pd.DataFrame:
    parquet_path = os.path.join(SCRIPT_DIR, nome_parquet)
    if os.path.exists(parquet_path):
        logger.info(f"  Cache encontrado → {nome_parquet}")
        return pd.read_parquet(parquet_path)

    logger.info(f"  Extraindo → {nome_parquet} ...")
    df = pd.read_sql(sql, conn)
    df.to_parquet(parquet_path, index=False)
    logger.info(f"  {len(df):,} linhas salvas em {nome_parquet}")
    return df


def extrair_todos(conn) -> dict:
    """
    Puxa cada tabela separada, sem JOIN.
    stores e addresses são pequenas — puxadas uma vez.
    sales_items e sales são puxadas por período.
    """
    dfs = {}

    # Tabelas de dimensão (sem filtro de data)
    dfs["stores"]    = extrair_tabela(conn, "raw_stores.parquet",
                           "SELECT id, name, address_id FROM stores")
    dfs["addresses"] = extrair_tabela(conn, "raw_addresses.parquet",
                           "SELECT id, state, city FROM addresses")

    # Fatos por período
    for nome, (dt_ini, dt_fim) in PERIODOS.items():
        dfs[f"items_{nome}"] = extrair_tabela(
            conn, f"raw_items_{nome}.parquet",
            f"SELECT id, sales_id, product_id, product_name, category_id, "
            f"category_name, quantity, price, created_at "
            f"FROM sales_items "
            f"WHERE created_at BETWEEN '{dt_ini} 00:00:00' AND '{dt_fim} 23:59:59'"
        )
        dfs[f"sales_{nome}"] = extrair_tabela(
            conn, f"raw_sales_{nome}.parquet",
            f"SELECT id, store_id FROM sales "
            f"WHERE created_at BETWEEN '{dt_ini} 00:00:00' AND '{dt_fim} 23:59:59' "
            f"AND status = 'paid'"
        )

    return dfs


def montar_periodo(dfs: dict, nome: str) -> pd.DataFrame:
    """Junta sales_items + sales + stores + addresses no pandas."""
    items  = dfs[f"items_{nome}"]
    sales  = dfs[f"sales_{nome}"][["id", "store_id"]].rename(columns={"id": "sales_id"})
    stores = dfs["stores"].rename(columns={"id": "store_id", "name": "loja", "address_id": "address_id"})
    addrs  = dfs["addresses"].rename(columns={"id": "address_id", "state": "estado", "city": "cidade"})

    df = (items
          .merge(sales,  on="sales_id",   how="inner")
          .merge(stores, on="store_id",   how="left")
          .merge(addrs,  on="address_id", how="left"))

    df["valor_rs"] = (df["price"] * df["quantity"] / 100.0).round(2)
    return df


# ---------------------------------------------------------------------------
# Filtro de chocolate (aplicado no pandas após extração)
# ---------------------------------------------------------------------------
def filtrar_chocolate(df: pd.DataFrame) -> pd.DataFrame:
    mask = pd.Series(False, index=df.index)
    for pat in CHOC_CATEGORIES:
        clean = pat.replace("%", "")
        mask |= df["category_name"].str.contains(clean, case=False, na=False)
        mask |= df["product_name"].str.contains(clean, case=False, na=False)
    return df[mask].copy()


# ---------------------------------------------------------------------------
# Lojas ativas nos dois períodos (equivalente ao INTERSECT do SQL)
# ---------------------------------------------------------------------------
def lojas_em_ambos(df_a: pd.DataFrame, df_b: pd.DataFrame) -> set:
    return set(df_a["store_id"].unique()) & set(df_b["store_id"].unique())


# ---------------------------------------------------------------------------
# Agregação por loja
# ---------------------------------------------------------------------------
def agregar_lojas(df: pd.DataFrame, lojas: set, periodo_label: str) -> pd.DataFrame:
    sub = df[df["store_id"].isin(lojas)].copy()
    agg = (
        sub.groupby(["loja", "estado", "cidade"], as_index=False)
        .agg(Quantidade=("quantity", "sum"), Valor_RS=("valor_rs", "sum"))
    )
    agg.rename(columns={"loja": "Loja", "estado": "Estado", "cidade": "Cidade"}, inplace=True)
    agg.insert(0, "Periodo", periodo_label)
    agg["Valor_RS"] = agg["Valor_RS"].round(2)
    return agg.sort_values(["Estado", "Cidade", "Loja"]).reset_index(drop=True)


# ---------------------------------------------------------------------------
# Formatação Excel
# ---------------------------------------------------------------------------
HEADER_FILL = PatternFill("solid", fgColor="1A1A2E")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
ALT_FILL    = PatternFill("solid", fgColor="F5F5F5")
BORDER      = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)

def format_sheet(ws, df: pd.DataFrame):
    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value     = col_name
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = BORDER

    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        fill = ALT_FILL if row_idx % 2 == 0 else PatternFill()
        for col_idx, value in enumerate(row, start=1):
            cell        = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = BORDER
            cell.fill   = fill
            col_name    = df.columns[col_idx - 1]
            if col_name == "Valor_RS":
                cell.number_format = "#,##0.00"
                cell.alignment     = Alignment(horizontal="right")
            elif col_name == "Quantidade":
                cell.number_format = "#,##0"
                cell.alignment     = Alignment(horizontal="right")

    for col_idx in range(1, len(df.columns) + 1):
        max_len = max(
            len(str(df.columns[col_idx - 1])),
            df.iloc[:, col_idx - 1].astype(str).str.len().max() if not df.empty else 0,
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 55)

    ws.freeze_panes = "A2"


def salvar_excel(sheets: dict, filepath: str):
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        for name, df in sheets.items():
            df.to_excel(writer, sheet_name=name[:31], index=False)
    buf.seek(0)
    wb = load_workbook(buf)
    for name, df in sheets.items():
        format_sheet(wb[name[:31]], df)
    wb.save(filepath)
    logger.info(f"  Salvo: {filepath}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    logger.info("=" * 55)
    logger.info("  InHouse — Chocolates Páscoa (MySQL → Parquet → Excel)")
    logger.info("=" * 55)

    hoje = date.today()
    conn = get_mysql_conn()

    try:
        # 1. Extração por tabela, sem JOIN (evita timeout do servidor)
        logger.info("[ EXTRAÇÃO — sem JOIN no MySQL ]")
        dfs = extrair_todos(conn)
    finally:
        conn.close()

    # 2. Montar períodos no pandas (merge local)
    logger.info("[ MERGE LOCAL ]")
    df_p26 = montar_periodo(dfs, "pascoa_2026")
    df_sem = montar_periodo(dfs, "semana_ant")
    df_p25 = montar_periodo(dfs, "pascoa_2025")

    # 2. Filtro chocolate (pandas, sem peso no MySQL)
    logger.info("[ FILTRO CHOCOLATE ]")
    choc_p26 = filtrar_chocolate(df_p26)
    choc_sem = filtrar_chocolate(df_sem)
    choc_p25 = filtrar_chocolate(df_p25)
    logger.info(f"  Páscoa 2026 : {len(choc_p26):,} itens | {choc_p26['store_id'].nunique()} lojas")
    logger.info(f"  Semana ant. : {len(choc_sem):,} itens | {choc_sem['store_id'].nunique()} lojas")
    logger.info(f"  Páscoa 2025 : {len(choc_p25):,} itens | {choc_p25['store_id'].nunique()} lojas")

    # 3. Lojas ativas nos dois períodos (INTERSECT)
    lojas_aba_a = lojas_em_ambos(choc_p26, choc_sem)
    lojas_aba_b = lojas_em_ambos(choc_p26, choc_p25)
    logger.info(f"  Lojas Aba A (2026 ∩ semana anterior): {len(lojas_aba_a)}")
    logger.info(f"  Lojas Aba B (2026 ∩ 2025):            {len(lojas_aba_b)}")

    # 4. Agregação
    logger.info("[ AGREGAÇÃO ]")
    aba_a = pd.concat([
        agregar_lojas(choc_p26, lojas_aba_a, "02/04 a 05/04/2026  — Páscoa 2026"),
        agregar_lojas(choc_sem, lojas_aba_a, "26/03 a 29/03/2026  — Semana anterior"),
    ], ignore_index=True)

    aba_b = pd.concat([
        agregar_lojas(choc_p26, lojas_aba_b, "02/04 a 05/04/2026  — Páscoa 2026"),
        agregar_lojas(choc_p25, lojas_aba_b, "02/04 a 05/04/2025  — Páscoa 2025"),
    ], ignore_index=True)

    # Totais
    for label, df_ in [("Aba A", aba_a), ("Aba B", aba_b)]:
        for per in df_["Periodo"].unique():
            sub = df_[df_["Periodo"] == per]
            logger.info(f"  {label} [{per.strip()}] → {sub['Quantidade'].sum():,.0f} un | R$ {sub['Valor_RS'].sum():,.2f}")

    # 5. Excel
    logger.info("[ EXCEL ]")
    arq = os.path.join(SCRIPT_DIR, f"Chocolates_Pascoa_MySQL_{hoje.strftime('%Y%m%d')}.xlsx")
    salvar_excel({
        "Pascoa_vs_Semana_Anterior": aba_a,
        "Pascoa_2026_vs_2025":       aba_b,
    }, arq)

    logger.info("=" * 55)
    logger.info(f"  Concluído → {os.path.basename(arq)}")
    logger.info("=" * 55)


if __name__ == "__main__":
    main()
